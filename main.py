from __future__ import print_function
from openpyxl import load_workbook
from fastapi import FastAPI
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware

import os.path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# The ID and range of a sample spreadsheet.
SAMPLE_SPREADSHEET_ID = '11nM2bIQvUjVlG8wEuvqESBkixjoPWjesInbevzDTTSg'

SAMPLE_RANGE_NAME = 'A2:D21'

class Entry(BaseModel):
    site: str
    name: str
    status: str
    date: str
    comment: str
    sheet: str



app = FastAPI()
origins = [
    "http://localhost",
    "http://localhost:8080",
    "http://localhost:8081",
    "http://localhost:8082",
    "https://excel.bryceeppler.com",
    "*"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)




@app.get("/")
async def root():
    return {"message": "Welcome to the API"}


@app.get("/worksheets")
async def worksheets():
    wb = load_workbook('form-submission-book.xlsx')
    sheets = wb.sheetnames
    wb.close()
    return wb.sheetnames


@app.post("/post-to-sheet")
async def post_to_sheet(entry : Entry):
    # load the excel file, form-submission-book.xlsx
    wb = load_workbook('form-submission-book.xlsx')

    # get the sheet name from the entry
    sheet_name = entry.sheet
    ws = wb[sheet_name]

    # get the next available row
    next_row = ws.max_row + 1

    # add the row
    ws[f'A{next_row}'] = entry.site
    ws[f'B{next_row}'] = entry.name
    ws[f'C{next_row}'] = entry.status
    ws[f'D{next_row}'] = entry.date
    ws[f'E{next_row}'] = entry.comment

    # save the excel file
    wb.save('form-submission-book.xlsx')

    # close the excel file
    wb.close()

    return entry

@app.get("/get-from-sheets")
async def get_from_sheets():
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('sheets', 'v4', credentials=creds)

        # Call the Sheets API
        sheet = service.spreadsheets()
        result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                    range=SAMPLE_RANGE_NAME).execute()
        values = result.get('values', [])

        if not values:
            print('No data found.')
            return

        print('Name, Site, Crew, Date:')
        for row in values:
            # Print columns A through D
            print('%s, %s, %s, %s' % (row[0], row[1], row[2], row[3]))
    except HttpError as err:
        print(err)
